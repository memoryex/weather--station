import requests
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
import os
import time

# --- CONFIGURATION ---
# ThingSpeak Read Keys
READ_KEYS = {
    "463450": "VAL3TD2W5LADX7K1", # PLXE 1-4
    "703669": "S44OBKWC5C7FODZ5", # NOBO 1-4
    "802414": "I6NIZAVZYLPVV1ME", # NOBO 5-7, PLXE 5
    "807602": "WUO1DG7GXYNZP6SG"  # QRAD, XLE, etc.
}

# Line definitions matching the dashboard and the requested Excel structure
# Note: LSTE is mapped to PLXE 1 (Channel 463450, Field 1) as inferred from general context
LINES = [
    {"name": "LSTE",      "channel": "463450", "fieldCount": 1, "fieldBarcode": 2, "color": "FFC6EFCE"},
    {"name": "PLXE 2",    "channel": "463450", "fieldCount": 3, "fieldBarcode": 4, "color": "FFFFCCFF"},
    {"name": "PLXE 3",    "channel": "463450", "fieldCount": 5, "fieldBarcode": 6, "color": "FFFFCCFF"},
    {"name": "PLXE 4",    "channel": "463450", "fieldCount": 7, "fieldBarcode": 8, "color": "FFFFCCFF"},
    {"name": "QRAD",      "channel": "807602", "fieldCount": 3, "fieldBarcode": 4, "color": "FFCCFFFF"},
    {"name": "NOBO 1",    "channel": "703669", "fieldCount": 1, "fieldBarcode": 2, "color": "FFE2EFDA"},
    {"name": "NOBO 2",    "channel": "703669", "fieldCount": 3, "fieldBarcode": 4, "color": "FFE2EFDA"},
    {"name": "NOBO 3",    "channel": "703669", "fieldCount": 5, "fieldBarcode": 6, "color": "FFE2EFDA"},
    {"name": "NOBO 4",    "channel": "703669", "fieldCount": 7, "fieldBarcode": 8, "color": "FFE2EFDA"},
]

# Work intervals for the day shift
INTERVALS = [
    ("06:00", "07:00"), ("07:00", "08:00"), ("08:00", "09:00"), ("09:00", "10:00"),
    ("10:00", "11:00"), ("11:30", "12:00"), ("12:00", "13:00"), ("13:00", "14:00"),
    ("14:10", "15:00"), ("15:00", "16:00")
]

EXCEL_FILE = "GD_Gamybos_Ataskaita.xlsx"

def fetch_ts_data(channel, field, start_dt, end_dt):
    """Fetches feeds from ThingSpeak for a given channel/field in a time window."""
    key = READ_KEYS.get(str(channel))
    url = f"https://api.thingspeak.com/channels/{channel}/fields/{field}.json"
    # ThingSpeak uses ISO8601. We specify timezone to match local floor time.
    params = {
        "api_key": key,
        "start": start_dt.strftime("%Y-%m-%dT%H:%M:%S"),
        "end": end_dt.strftime("%Y-%m-%dT%H:%M:%S"),
        "timezone": "Europe/Vilnius"
    }
    try:
        r = requests.get(url, params=params, timeout=15)
        if r.status_code == 200:
            return r.json().get("feeds", [])
    except Exception as e:
        print(f"Error fetching data for channel {channel}: {e}")
    return []

def calculate_delta(feeds, field_name):
    """Calculates total production by summing positive counter increments."""
    if not feeds: return 0
    vals = []
    for f in feeds:
        v = f.get(field_name)
        if v is not None and v != "null" and v != "":
            try:
                vals.append(float(v))
            except ValueError:
                continue

    if not vals: return 0
    if len(vals) == 1: return 0 # Need at least two points to see a change

    total = 0
    for i in range(1, len(vals)):
        diff = vals[i] - vals[i-1]
        if diff > 0:
            total += diff
        # If diff is negative, it's a reset; we assume the current value is the production since reset
        # but to be safe and match dashboard logic, we only count positive jumps.
    return int(total)

def get_last_val(feeds, field_name):
    """Gets the last non-null value from a list of feeds."""
    for f in reversed(feeds):
        v = f.get(field_name)
        if v is not None and v != "null" and v != "":
            return str(v)
    return ""

def apply_styles(ws):
    """Applies basic styling to the worksheet."""
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_column=13):
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = PatternFill(start_color="FF333333", end_color="FF333333", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

def setup_sheet(ws, date_str):
    """Sets up the initial structure of a new daily sheet."""
    headers = ["Linija", "Laukas", "6:00-7:00", "7:00-8:00", "8:00-9:00", "9:10-10:00", "10:00-11:00", "11:30-12:00", "12:00-13:00", "13:00-14:00", "14:10-15:00", "15:00-16:00", "Viso"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c).value = h

    current_row = 2
    for line in LINES:
        name = line["name"]
        ws.cell(row=current_row, column=1).value = name
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+3, end_column=1)
        ws.cell(row=current_row, column=1).alignment = Alignment(vertical="center", horizontal="center", text_rotation=90)
        ws.cell(row=current_row, column=1).font = Font(bold=True)

        ws.cell(row=current_row,     column=2).value = f"{name} Planas"
        ws.cell(row=current_row + 1, column=2).value = f"{name} Faktas"
        ws.cell(row=current_row + 2, column=2).value = "Gaminys"
        ws.cell(row=current_row + 3, column=2).value = "Komentaras"

        # Colors
        fill = PatternFill(start_color=line["color"], end_color=line["color"], fill_type="solid")
        for r_offset in range(4):
            ws.cell(row=current_row + r_offset, column=2).fill = fill

        # Fact Sum Formula
        col_letters = "CDEFGHIJKL" # 10 columns
        sum_formula = f"=SUM(C{current_row+1}:L{current_row+1})"
        ws.cell(row=current_row + 1, column=13).value = sum_formula
        ws.cell(row=current_row + 1, column=13).font = Font(bold=True)

        current_row += 4

    apply_styles(ws)

def run_update():
    """Main function to update Excel with data from ThingSpeak."""
    now = datetime.datetime.now()
    date_str = now.strftime("%Y-%m-%d")

    if os.path.exists(EXCEL_FILE):
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
        except Exception:
            print("Could not open existing Excel file. It might be open in another program.")
            return
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames: del wb["Sheet"]

    if date_str in wb.sheetnames:
        ws = wb[date_str]
    else:
        ws = wb.create_sheet(date_str)
        setup_sheet(ws, date_str)

    # Sheet protection
    ws.protection.sheet = True
    ws.protection.password = "gd2024"

    print(f"Updating data for {date_str}...")

    line_row = 2
    for line in LINES:
        print(f"  Processing {line['name']}...")
        field_count = f"field{line['fieldCount']}"
        field_barcode = f"field{line['fieldBarcode']}"

        for idx, (start_t, end_t) in enumerate(INTERVALS):
            col = 3 + idx

            # Define time window for this cell
            start_dt = datetime.datetime.strptime(f"{date_str} {start_t}", "%Y-%m-%d %H:%M")
            end_dt = datetime.datetime.strptime(f"{date_str} {end_t}", "%Y-%m-%d %H:%M")

            # Skip future intervals
            if start_dt > now:
                continue

            # Fetch and calculate production
            feeds = fetch_ts_data(line["channel"], line["fieldCount"], start_dt, end_dt)
            produced = calculate_delta(feeds, field_count)

            # Update Fact Cell
            fact_cell = ws.cell(row=line_row + 1, column=col)
            fact_cell.value = produced
            fact_cell.protection = Protection(locked=True)
            fact_cell.font = Font(bold=True)

            # Fetch and update Barcode
            b_feeds = fetch_ts_data(line["channel"], line["fieldBarcode"], start_dt, end_dt)
            barcode = get_last_val(b_feeds, field_barcode)

            prod_cell = ws.cell(row=line_row + 2, column=col)
            if barcode:
                prod_cell.value = f"X-{barcode}"
            prod_cell.protection = Protection(locked=False) # Allow manual override if needed

            # Ensure Plan and Comment are unlocked
            ws.cell(row=line_row, column=col).protection = Protection(locked=False)
            ws.cell(row=line_row + 3, column=col).protection = Protection(locked=False)

        line_row += 4

    try:
        wb.save(EXCEL_FILE)
        print(f"Successfully saved to {EXCEL_FILE}")
    except Exception as e:
        print(f"Error saving file: {e}. Make sure the file is not open in Excel.")

if __name__ == "__main__":
    run_update()
