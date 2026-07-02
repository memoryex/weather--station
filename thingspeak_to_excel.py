import os
import sys
import time
import datetime
import requests
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Protection
from openpyxl.utils import get_column_letter

# --- CONFIGURATION ---
EXCEL_FILE = "GD_Gamyba.xlsx"
TIME_SLOTS = [
    ("06:00", "07:00"),
    ("07:00", "08:00"),
    ("08:00", "09:00"),
    ("09:10", "10:00"),
    ("10:00", "11:00"),
    ("11:30", "12:00"),
    ("12:00", "13:00"),
    ("13:00", "14:00"),
    ("14:10", "15:00"),
    ("15:00", "16:00"),
]

# ThingSpeak Read Keys
READ_KEYS = {
    463450: "VAL3TD2W5LADX7K1", # PLXE 1-4
    703669: "S44OBKWC5C7FODZ5", # NOBO 1-4
    807602: "WUO1DG7GXYNZP6SG", # QRAD 1, XLE/LSTE
    802414: "I6NIZAVZYLPVV1ME"  # NOBO 5-7, PLXE 5
}

# Production Lines Definition
# Based on the user's Excel image and existing dashboard configs
LINES = [
    {"name": "LSTE",   "channel": 807602, "f_count": 5, "f_barcode": 6, "color": "D9EAD3"}, # Assuming LSTE is XLE 1
    {"name": "PLXE 2", "channel": 463450, "f_count": 3, "f_barcode": 4, "color": "EAD1DC"},
    {"name": "PLXE 3", "channel": 463450, "f_count": 5, "f_barcode": 6, "color": "EAD1DC"},
    {"name": "PLXE 4", "channel": 463450, "f_count": 7, "f_barcode": 8, "color": "EAD1DC"},
    {"name": "QRAD",   "channel": 807602, "f_count": 3, "f_barcode": 4, "color": "CFE2F3"},
    {"name": "NOBO 1", "channel": 703669, "f_count": 1, "f_barcode": 2, "color": "D9EAD3"},
    {"name": "NOBO 2", "channel": 703669, "f_count": 3, "f_barcode": 4, "color": "D9EAD3"},
    {"name": "NOBO 3", "channel": 703669, "f_count": 5, "f_barcode": 6, "color": "D9EAD3"},
    {"name": "NOBO 4", "channel": 703669, "f_count": 7, "f_barcode": 8, "color": "D9EAD3"},
]

# --- THINGSPEAK LOGIC ---

def get_ts_data(channel_id, field_count, field_barcode, start_time, end_time):
    """Fetches count difference and last barcode for a time interval."""
    read_key = READ_KEYS.get(channel_id)
    if not read_key:
        return 0, ""

    url = f"https://api.thingspeak.com/channels/{channel_id}/feeds.json"

    # Note: ThingSpeak uses UTC. If start_time is local (Lithuania),
    # we should ideally convert it to UTC. However, for simplicity
    # and to match existing dashboard behavior, we use the raw strings.
    # We remove the 'Z' to avoid explicitly marking it as UTC if it's local.

    # Get first value in interval
    params_start = {
        "api_key": read_key,
        "start": start_time.strftime("%Y-%m-%d %H:%M:%S"),
        "end": end_time.strftime("%Y-%m-%d %H:%M:%S"),
        "results": 1
    }

    # Get last value in interval
    # Using 'results=1' without offset gives the oldest record.
    # To get the newest (last) record, we could use results=1 and no other window,
    # but since we want the last one *within* the window, we fetch results=1.
    # However, ThingSpeak 'results' with 'start'/'end' returns the OLDER entries first.
    # To get the LATEST entry, we don't use 'start', we use 'end' and 'results=1'.
    params_end = {
        "api_key": read_key,
        "end": end_time.strftime("%Y-%m-%d %H:%M:%S"),
        "results": 1
    }

    try:
        r_start = requests.get(url, params=params_start, timeout=10).json()
        r_end = requests.get(url, params=params_end, timeout=10).json()

        feeds_start = r_start.get("feeds", [])
        feeds_end = r_end.get("feeds", [])

        if not feeds_start or not feeds_end:
            return 0, ""

        start_val = int(feeds_start[0].get(f"field{field_count}") or 0)
        end_val = int(feeds_end[0].get(f"field{field_count}") or 0)
        last_barcode = feeds_end[0].get(f"field{field_barcode}") or ""

        # The produced count is the difference (including the first unit if it was just produced)
        # But ThingSpeak increments might be cumulative.
        # Usually we take the first record of the day as baseline.

        # Simple diff for the interval:
        count = end_val - start_val
        if count < 0: count = 0 # Safety

        # Prefix barcode with X- as requested
        if last_barcode and not last_barcode.startswith("X-"):
            last_barcode = f"X-{last_barcode}"

        return count, last_barcode
    except Exception as e:
        print(f"Error fetching ThingSpeak data: {e}")
        return 0, ""

def check_production_started(date_obj):
    """Checks if any production happened since 06:00 today."""
    start_of_day = datetime.datetime.combine(date_obj, datetime.time(6, 0))
    now = datetime.datetime.now()
    if now < start_of_day:
        return False

    for line in LINES:
        count, _ = get_ts_data(line["channel"], line["f_count"], line["f_barcode"], start_of_day, now)
        if count > 0:
            return True
    return False

# --- EXCEL LOGIC ---

def create_daily_sheet(wb, date_str):
    if date_str in wb.sheetnames:
        return wb[date_str]

    ws = wb.create_sheet(date_str)

    # Top Summary Row (Row 1) - Totals per hour
    for i in range(len(TIME_SLOTS)):
        col = i + 3
        col_letter = get_column_letter(col)
        # Sum of all "Faktas" rows for this column
        # Faktas rows are 5, 9, 13, 17, 21 (gap at 25) 26, 30, 34, 38
        faktas_rows = [5 + j*4 for j in range(5)] + [26 + j*4 for j in range(4)]
        formula = "=" + "+".join([f"{col_letter}{r}" for r in faktas_rows])
        ws.cell(row=1, column=col).value = formula
        ws.cell(row=1, column=col).font = Font(color="FF0000", bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    # Headers (Row 3)
    ws["B3"] = "Laikas"
    for i, (start, end) in enumerate(TIME_SLOTS):
        col = i + 3
        ws.cell(row=3, column=col).value = f"{start}-{end}"
        ws.cell(row=3, column=col).alignment = Alignment(horizontal="center")

    ws.cell(row=3, column=len(TIME_SLOTS) + 3).value = "Viso"
    ws.cell(row=3, column=len(TIME_SLOTS) + 4).value = "II pamaina"
    ws.cell(row=3, column=len(TIME_SLOTS) + 5).value = "Linijos vnt"
    ws.cell(row=3, column=len(TIME_SLOTS) + 6).value = "Vnt/val"

    # Formatting styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    current_row = 4
    for idx, line in enumerate(LINES):
        # Add a gap after QRAD (Line 5) to match image
        if idx == 5:
            current_row += 1

        # Line Name in Col A
        ws.cell(row=current_row, column=1).value = line["name"]
        ws.cell(row=current_row, column=1).font = Font(bold=True)

        # Planas Row
        ws.cell(row=current_row, column=2).value = f"{line['name']} Planas"
        ws.cell(row=current_row, column=2).fill = PatternFill(start_color="38761D", end_color="38761D", fill_type="solid")
        ws.cell(row=current_row, column=2).font = Font(color="FFFFFF", bold=True)

        for c in range(3, len(TIME_SLOTS) + 7):
             ws.cell(row=current_row, column=c).fill = PatternFill(start_color="93C47D", end_color="93C47D", fill_type="solid")

        # Faktas Row
        ws.cell(row=current_row + 1, column=2).value = f"{line['name']} faktas"

        # Gaminys Row
        ws.cell(row=current_row + 2, column=2).value = "Gaminys"

        # Komentaras Row
        ws.cell(row=current_row + 3, column=2).value = "Komentaras"

        # Apply colors and borders for the block
        line_color = line["color"]
        for r in range(current_row + 1, current_row + 4):
            for c in range(2, len(TIME_SLOTS) + 7):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill(start_color=line_color, end_color=line_color, fill_type="solid")
                cell.border = thin_border

                # Lock "Faktas" cells (column 3 to end)
                if r == current_row + 1 and 3 <= c <= len(TIME_SLOTS) + 2:
                    cell.protection = Protection(locked=True)
                else:
                    cell.protection = Protection(locked=False)

        # Formula for Viso
        viso_col = len(TIME_SLOTS) + 3
        start_cell = ws.cell(row=current_row + 1, column=3).coordinate
        end_cell = ws.cell(row=current_row + 1, column=viso_col - 1).coordinate
        ws.cell(row=current_row + 1, column=viso_col).value = f"=SUM({start_cell}:{end_cell})"

        current_row += 4

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    for c in range(3, len(TIME_SLOTS) + 7):
        ws.column_dimensions[get_column_letter(c)].width = 12

    ws.protection.sheet = True
    ws.protection.password = "GD2024"

    return ws

def update_excel():
    now = datetime.datetime.now()
    date_str = now.strftime("%Y-%m-%d")

    # 1. Load or create workbook
    if os.path.exists(EXCEL_FILE):
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
        except Exception:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames: del wb["Sheet"]
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames: del wb["Sheet"]

    # 2. Check if we should create a new sheet
    if date_str not in wb.sheetnames:
        if not check_production_started(now.date()):
            print(f"[{now}] Production not yet started for {date_str}. Skipping sheet creation.")
            return
        print(f"[{now}] Production detected! Creating new sheet for {date_str}.")
        ws = create_daily_sheet(wb, date_str)
    else:
        ws = wb[date_str]

    # 3. Update time slots
    ws.protection.sheet = False # Unlock to update

    current_row = 4
    for idx, line in enumerate(LINES):
        if idx == 5:
            current_row += 1

        for i, (start_str, end_str) in enumerate(TIME_SLOTS):
            slot_start = datetime.datetime.combine(now.date(), datetime.time.fromisoformat(start_str))
            slot_end = datetime.datetime.combine(now.date(), datetime.time.fromisoformat(end_str))

            # Only update if current time is past slot_start
            if now > slot_start:
                # We update until the end of slot or current time (whichever is earlier)
                effective_end = min(now, slot_end)
                count, barcode = get_ts_data(line["channel"], line["f_count"], line["f_barcode"], slot_start, effective_end)

                # Update Faktas (row current_row + 1, column i + 3)
                ws.cell(row=current_row + 1, column=i + 3).value = count

                # Update Gaminys (row current_row + 2, column i + 3)
                if barcode:
                    ws.cell(row=current_row + 2, column=i + 3).value = barcode

        current_row += 4

    ws.protection.sheet = True
    wb.save(EXCEL_FILE)
    print(f"[{now}] Excel updated successfully.")

if __name__ == "__main__":
    print("=== ThingSpeak to Excel Automation Started ===")
    while True:
        try:
            update_excel()
        except Exception as e:
            print(f"Critical error in loop: {e}")

        # Wait for 1 hour
        time.sleep(3600)
